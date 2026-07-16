"""
Import stops, routes, and route_stops from Sajha_Routes_Cleaned.xlsx into PostgreSQL/PostGIS.

Deduplication strategy:
  - Round (latitude, longitude) to 5 decimal places.
  - Each unique rounded coordinate pair = one row in `stops`.
  - Stop ID is auto-generated as "S{index}".

Operator parsing:
  - The operator name is extracted from the Route_ID prefix (everything
    before the first hyphen, e.g. "Sajha" from "Sajha-101").

Usage:
  cd backend
  python -m scripts.import_stops
"""

import sys
import os
import re
import logging

# Ensure the backend directory is on sys.path so `app` package is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy import text
from app.database import engine, SessionLocal, Base
from app.models import Stop, Route, RouteStop

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "Sajha_Routes_Cleaned.xlsx")


def ensure_postgis(conn):
    """Enable PostGIS extension if not already enabled."""
    conn.execute(text("CREATE EXTENSION IF NOT EXISTS postgis"))
    conn.commit()
    logger.info("PostGIS extension ensured.")


def parse_operator(route_id: str) -> str:
    """Extract operator name from Route_ID prefix (before first hyphen or digit sequence)."""
    # Examples: "Sajha-101" -> "Sajha", "Mahanagar-5" -> "Mahanagar"
    match = re.match(r"^([A-Za-z\s]+)", route_id)
    if match:
        return match.group(1).strip()
    return route_id


def read_xlsx(path: str) -> list[dict]:
    """Read the xlsx file and return a list of row dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info(f"Excel headers: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        # Skip rows with missing critical data
        if row_dict.get("Route_ID") and row_dict.get("latitude") is not None and row_dict.get("longitude") is not None:
            rows.append(row_dict)

    wb.close()
    logger.info(f"Read {len(rows)} valid data rows from xlsx.")
    return rows


def import_data():
    """Main import function."""
    # ---- Read xlsx ----
    rows = read_xlsx(XLSX_PATH)
    if not rows:
        logger.error("No data rows found. Aborting.")
        return

    # ---- Enable PostGIS ----
    with engine.connect() as conn:
        ensure_postgis(conn)

    # ---- Clean slate: drop everything and recreate ----
    with engine.connect() as conn:
        conn.execute(text("DROP INDEX IF EXISTS idx_stops_geom"))
        conn.execute(text("DROP INDEX IF EXISTS idx_stops_geom_geom"))  # GeoAlchemy2 auto-name
        conn.commit()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    logger.info("Tables created (dropped old ones first).")

    # ---- Deduplicate stops by rounded coordinates ----
    # Key: (rounded_lat, rounded_lng) -> stop info
    coord_to_stop = {}
    # Track stop names per coordinate (pick first occurrence)
    stop_counter = 0

    for row in rows:
        lat = round(float(row["latitude"]), 5)
        lng = round(float(row["longitude"]), 5)
        key = (lat, lng)

        if key not in coord_to_stop:
            stop_counter += 1
            stop_name = row.get("Stop_Name", f"Stop {stop_counter}")
            coord_to_stop[key] = {
                "id": f"S{stop_counter}",
                "name": stop_name,
                "latitude": lat,
                "longitude": lng,
            }

    logger.info(f"Deduplicated to {len(coord_to_stop)} unique stops.")

    # ---- Build routes ----
    route_ids_seen = set()
    routes_data = []

    for row in rows:
        rid = str(row["Route_ID"]).strip()
        if rid not in route_ids_seen:
            route_ids_seen.add(rid)
            operator = parse_operator(rid)
            routes_data.append({
                "id": rid,
                "name": rid,            # use Route_ID as name
                "operator": operator,
            })

    logger.info(f"Found {len(routes_data)} unique routes.")

    # ---- Build route_stops ----
    # Map (rounded_lat, rounded_lng) -> stop_id for fast lookup
    coord_to_id = {k: v["id"] for k, v in coord_to_stop.items()}

    route_stops_data = []
    for row in rows:
        rid = str(row["Route_ID"]).strip()
        lat = round(float(row["latitude"]), 5)
        lng = round(float(row["longitude"]), 5)
        stop_id = coord_to_id[(lat, lng)]
        stop_order = int(row["Stop_Sequence"])

        route_stops_data.append({
            "route_id": rid,
            "stop_id": stop_id,
            "stop_order": stop_order,
        })

    logger.info(f"Built {len(route_stops_data)} route_stop entries.")

    # ---- Insert into DB ----
    db = SessionLocal()
    try:
        # Insert stops
        stop_objects = []
        for sdata in coord_to_stop.values():
            stop_objects.append(Stop(
                id=sdata["id"],
                name=sdata["name"],
                latitude=sdata["latitude"],
                longitude=sdata["longitude"],
                geom=f"SRID=4326;POINT({sdata['longitude']} {sdata['latitude']})",
            ))
        db.add_all(stop_objects)
        db.flush()
        logger.info(f"Inserted {len(stop_objects)} stops.")

        # Insert routes
        route_objects = []
        for rdata in routes_data:
            route_objects.append(Route(
                id=rdata["id"],
                name=rdata["name"],
                operator=rdata["operator"],
                vehicle_type="bus",
            ))
        db.add_all(route_objects)
        db.flush()
        logger.info(f"Inserted {len(route_objects)} routes.")

        # Insert route_stops
        rs_objects = []
        for rsdata in route_stops_data:
            rs_objects.append(RouteStop(
                route_id=rsdata["route_id"],
                stop_id=rsdata["stop_id"],
                stop_order=rsdata["stop_order"],
            ))
        db.add_all(rs_objects)
        db.flush()
        logger.info(f"Inserted {len(rs_objects)} route_stops.")

        db.commit()
        logger.info("All data committed successfully.")

        # ---- Report counts ----
        stop_count = db.query(Stop).count()
        route_count = db.query(Route).count()
        rs_count = db.query(RouteStop).count()

        logger.info("=" * 50)
        logger.info(f"  stops:       {stop_count}")
        logger.info(f"  routes:      {route_count}")
        logger.info(f"  route_stops: {rs_count}")
        logger.info("=" * 50)

    except Exception as e:
        db.rollback()
        logger.error(f"Error during import: {e}", exc_info=True)
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import_data()
